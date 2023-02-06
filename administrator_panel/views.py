import datetime
import locale
from copy import deepcopy
import pandas as pd
import pdfkit

import openpyxl
from django.contrib import messages
from django.contrib.auth import update_session_auth_hash
from django.core.mail import EmailMessage
from django.views.generic.base import ContextMixin, TemplateResponseMixin
from openpyxl.styles import Alignment
from openpyxl.utils import get_column_letter

from django.db.models import Q, Sum, Value, Case, When, FloatField, Subquery, OuterRef, QuerySet, F
from django.db.models.functions import Concat, TruncMonth
from django.http import Http404, JsonResponse
from django.shortcuts import redirect
from django.urls import reverse
from django.views import View
from django.db.models.deletion import ProtectedError
from django.views.generic.edit import FormMixin
from django.views.generic.detail import SingleObjectMixin
from django.views.generic.list import MultipleObjectMixin

from configuration.models import Role, Tariff, Service, MeasurementUnit, PaymentRequisite, ArticlePayment
from .forms import *
from .mixins import *
from .functions import owner_context_data, calculate_notoriety_and_receipt_sum, count_all_totals, calculate_totals

from my_house_24 import settings


locale.setlocale(locale.LC_ALL, 'uk_UA.utf8')   # for using stringify('%B') with datetime objects


class StatisticListView(PermissionListView):
    model = Receipt
    template_name = 'administrator_panel/statistics-list.html'
    string_permission = 'statistic_access'

    def get_queryset(self):
        return None

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=object_list, **kwargs)
        context['house_sum'] = House.objects.all().count()
        context['active_owner_sum'] = User.objects.filter(role__role='owner').count()
        context['application_in_work_sum'] = Application.objects.filter(status='in work').count()
        context['flat_sum'] = Flat.objects.all().count()

        personal_accounts = PersonalAccount.objects.all()
        context['account_sum'] = personal_accounts.count()
        context['application_new_sum'] = Application.objects.filter(status='new').count()

        context['totals'] = count_all_totals(personal_accounts)
        notorieties = Notoriety.objects.filter(created_at__range=(timezone.datetime(timezone.datetime.now().year - 1,
                                                                                    timezone.datetime.now().month,
                                                                                    timezone.datetime.now().day),
                                                                  timezone.datetime.now()))
        receipts = Receipt.objects\
            .prefetch_related('receiptservices')\
            .filter(is_completed=True, date_from__range=(timezone.datetime(timezone.datetime.now().year - 1,
                                                                           timezone.datetime.now().month,
                                                                           timezone.datetime.now().day),
                                                         timezone.datetime.now()))\
            .order_by('date_from')

        grouped_income_notorieties_queryset = notorieties\
            .filter(type='income')\
            .annotate(month=TruncMonth('created_at'))\
            .values('month')\
            .annotate(sum=Sum('sum'))
        grouped_outcome_notorieties_queryset = notorieties\
            .filter(type='outcome') \
            .annotate(month=TruncMonth('created_at'))\
            .values('month')\
            .annotate(sum=Sum('sum'))

        grouped_receipts_queryset = receipts\
            .annotate(month=TruncMonth('date_from'))\
            .values('month')\
            .annotate(sum=Sum('receiptservices__total_price'))

        # forming dictionary with outcome through all year (it includes notorieties with 'outcome' status and completed receipts)
        context['outcome_sum'] = {}
        for obj in grouped_receipts_queryset:
            if not context['outcome_sum'].get(obj.get('month')):
                context['outcome_sum'][obj.get('month')] = obj.get('sum') if obj.get('sum') else 0.00
            else:
                context['outcome_sum'][obj.get('month')] += obj.get('sum') if obj.get('sum') else 0.00

        # forming dictionary with income through all year (it includes only notorieties with 'income' status)
        context['income_sum'] = {}
        for obj in grouped_income_notorieties_queryset:
            if not context['income_sum'].get(obj.get('month')):
                context['income_sum'][obj.get('month')] = obj.get('sum') if obj.get('sum') else 0.00
            else:
                context['income_sum'][obj.get('month')] += obj.get('sum') if obj.get('sum') else 0.00

        context['income_sum_for_notorieties'] = deepcopy(context['income_sum'])     # copy other one income sum for display in chart with notorieties

        # making keys of each dictionary equal in order to correct display datasets in chart
        for obj in context['income_sum']:
            if not context['outcome_sum'].get(obj):
                context['outcome_sum'][obj] = 0.00

        for obj in context['outcome_sum']:
            if not context['income_sum'].get(obj):
                context['income_sum'][obj] = 0.00

        context['income_sum'] = dict(sorted(context['income_sum'].items()))
        context['outcome_sum'] = dict(sorted(context['outcome_sum'].items()))

        # replacing datetime objects to their month display
        for key, value in list(context['income_sum'].items()):
            context['income_sum'][key.strftime('%B').title()] = context['income_sum'].pop(key)

        for key, value in list(context['outcome_sum'].items()):
            context['outcome_sum'][key.strftime('%B').title()] = context['outcome_sum'].pop(key)

        # forming outcome only with notorieties
        context['outcome_notorieties_sum'] = {}
        for obj in grouped_outcome_notorieties_queryset:
            if not context['outcome_notorieties_sum'].get(obj.get('month')):
                context['outcome_notorieties_sum'][obj.get('month')] = obj.get('sum') if obj.get('sum') else 0.00
            else:
                context['outcome_notorieties_sum'][obj.get('month')] += obj.get('sum') if obj.get('sum') else 0.00

        # making keys of each dictionary equal in order to correct display datasets in chart
        for obj in context['income_sum_for_notorieties']:
            if not context['outcome_notorieties_sum'].get(obj):
                context['outcome_notorieties_sum'][obj] = 0.00

        for obj in context['outcome_notorieties_sum']:
            if not context['income_sum_for_notorieties'].get(obj):
                context['income_sum_for_notorieties'][obj] = 0.00

        context['income_sum_for_notorieties'] = dict(sorted(context['income_sum_for_notorieties'].items()))
        context['outcome_notorieties_sum'] = dict(sorted(context['outcome_notorieties_sum'].items()))

        # replacing datetime objects to their month display
        for key, value in list(context['income_sum_for_notorieties'].items()):
            context['income_sum_for_notorieties'][key.strftime('%B').title()] = context['income_sum_for_notorieties'].pop(key)

        for key, value in list(context['outcome_notorieties_sum'].items()):
            context['outcome_notorieties_sum'][key.strftime('%B').title()] = context['outcome_notorieties_sum'].pop(key)

        return context


class HouseCreateView(PermissionCreateView):
    model = House
    template_name = 'administrator_panel/house-create-update.html'
    form_class = HouseForm
    string_permission = 'building_access'

    def get_context_data(self, **kwargs):
        context = super(HouseCreateView, self).get_context_data(**kwargs)
        context['user_formset'] = house_user_formset(queryset=HouseUser.objects.none(), prefix='users')
        context['section_formset'] = section_formset(queryset=Section.objects.none(), prefix='sections')
        context['floor_formset'] = floor_formset(queryset=Floor.objects.none(), prefix='floors')
        context['users'] = User.objects.select_related('role').all()

        # dictionary for parsing js dictionary on frontend for changeUser function
        context['user_role_dict'] = {'none': ''}  # starter value
        for user in context['users']:
            context['user_role_dict'][user.pk] = user.role.get_role_display()
        context['roles'] = Role.objects.all()
        return context

    def post(self, request, *args, **kwargs):
        form = HouseForm(request.POST, request.FILES or None)
        user_formset_post = house_user_formset(request.POST, queryset=HouseUser.objects.none(), prefix='users')
        section_formset_post = section_formset(request.POST, queryset=Section.objects.none(), prefix='sections')
        floor_formset_post = floor_formset(request.POST, queryset=Floor.objects.none(), prefix='floors')

        if form.is_valid() and user_formset_post.is_valid() and section_formset_post.is_valid() \
                and floor_formset_post.is_valid():
            house = self.form_valid(form=form,
                                    user_formset_post=user_formset_post,
                                    section_formset_post=section_formset_post,
                                    floor_formset_post=floor_formset_post)
            return redirect('house-detail', house_pk=house.pk)

        self.object = None
        context = self.get_context_data()
        context['form'] = form
        context['user_formset'] = user_formset_post
        context['floor_formset'] = floor_formset_post
        context['section_formset'] = section_formset_post
        return self.render_to_response(context)

    def form_valid(self, form, **kwargs):
        house = form.save()  # is used for selecting foreign key of new form in formsets

        for form in kwargs['user_formset_post']:
            if form.cleaned_data.get('user'):
                user_house_saved = form.save(commit=False)
                user_house_saved.house = house
                user_house_saved.save()

        for form in kwargs['section_formset_post']:
            if form.cleaned_data.get('name'):
                section_saved = form.save(commit=False)
                section_saved.house = house
                section_saved.save()

        for form in kwargs['floor_formset_post']:
            if form.cleaned_data.get('name'):
                floor_saved = form.save(commit=False)
                floor_saved.house = house
                floor_saved.save()

        return house


class HouseUpdateView(PermissionUpdateView):
    model = House
    template_name = 'administrator_panel/house-create-update.html'
    pk_url_kwarg = 'house_pk'
    form_class = HouseForm
    string_permission = 'building_access'

    def get_object(self, queryset=None):
        try:
            return House.objects.prefetch_related('houseuser_set', 'houseuser_set__user', 'houseuser_set__user__role',
                                                  'floor_set', 'section_set').get(pk=self.kwargs['house_pk'])
        except (House.DoesNotExist, ValueError, AttributeError):
            raise Http404()

    def get_context_data(self, **kwargs):
        self.object = self.get_object()
        context = super(HouseUpdateView, self).get_context_data(**kwargs)
        context['user_formset'] = house_user_formset(queryset=self.object.houseuser_set.all(), prefix='users')
        context['section_formset'] = section_formset(queryset=self.object.section_set.all(), prefix='sections')
        context['floor_formset'] = floor_formset(queryset=self.object.floor_set.all(), prefix='floors')
        context['users'] = User.objects.select_related('role').all()

        # dictionary for parsing js dictionary on frontend for changeUser function
        context['user_role_dict'] = {'none': ''}  # starter value
        for user in context['users']:
            context['user_role_dict'][user.pk] = user.role.get_role_display()
        context['roles'] = Role.objects.all()
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = HouseForm(request.POST, request.FILES or None, instance=self.object)
        user_formset_post = house_user_formset(request.POST, queryset=self.object.houseuser_set.all(), prefix='users')
        section_formset_post = section_formset(request.POST, queryset=self.object.section_set.all(), prefix='sections')
        floor_formset_post = floor_formset(request.POST, queryset=self.object.floor_set.all(), prefix='floors')

        if form.is_valid() and user_formset_post.is_valid() and section_formset_post.is_valid() \
                and floor_formset_post.is_valid():
            self.form_valid(form=form,
                            user_formset_post=user_formset_post,
                            section_formset_post=section_formset_post,
                            floor_formset_post=floor_formset_post)
            return redirect('house-detail', house_pk=self.kwargs['house_pk'])

        self.object = None
        context = self.get_context_data()
        context['form'] = form
        context['user_formset'] = user_formset_post
        context['floor_formset'] = floor_formset_post
        context['section_formset'] = section_formset_post
        return self.render_to_response(context)

    def form_valid(self, form, **kwargs):
        house = form.save()  # is used for selecting foreign key of new form in formsets

        for form in kwargs['user_formset_post']:
            if form.cleaned_data.get('user'):
                user_house_saved = form.save(commit=False)
                user_house_saved.house = house
                user_house_saved.save()

        for form in kwargs['section_formset_post']:
            if form.cleaned_data.get('name'):
                section_saved = form.save(commit=False)
                section_saved.house = house
                section_saved.save()

        for form in kwargs['floor_formset_post']:
            if form.cleaned_data.get('name'):
                floor_saved = form.save(commit=False)
                floor_saved.house = house
                floor_saved.save()


class HouseDetailView(PermissionDetailView):
    model = House
    template_name = 'administrator_panel/house-detail.html'
    pk_url_kwarg = 'house_pk'
    string_permission = 'building_access'

    def get_object(self, queryset=None):
        try:
            return House.objects.prefetch_related('houseuser_set', 'houseuser_set__user', 'houseuser_set__user__role',
                                                  'floor_set', 'section_set').get(pk=self.kwargs['house_pk'])
        except (House.DoesNotExist, ValueError, AttributeError):
            raise Http404()


class HouseListView(PermissionListView):
    model = House
    queryset = House.objects.all()
    template_name = 'administrator_panel/house-list.html'
    string_permission = 'building_access'


def delete_house(request, house_pk):
    if request.user.is_anonymous or not request.user.role.building_access:
        return JsonResponse({'answer': 'Ви не маєте доступу до будинків'})
    try:
        house = House.objects.get(pk=house_pk)
        house.delete()
        return JsonResponse({'answer': 'success'})
    except (ValueError, AttributeError, House.DoesNotExist, ProtectedError):
        return JsonResponse({'answer': 'failed'})


def delete_section(request, section_pk):
    if request.user.is_anonymous or not request.user.role.building_access:
        return JsonResponse({'answer': 'Ви не маєте доступу до будинків'})
    try:
        section_to_delete = Section.objects.get(pk=section_pk)
        section_to_delete.delete()
        return JsonResponse({'answer': 'success'})
    except (ValueError, AttributeError, Section.DoesNotExist):
        return JsonResponse({'answer': 'failed'})


def delete_floor(request, floor_pk):
    if request.user.is_anonymous or not request.user.role.building_access:
        return JsonResponse({'answer': 'Ви не маєте доступу до будинків'})
    try:
        floor_to_delete = Floor.objects.get(pk=floor_pk)
        floor_to_delete.delete()
        return JsonResponse({'answer': 'success'})
    except (ValueError, AttributeError, Floor.DoesNotExist):
        return JsonResponse({'answer': 'failed'})


def delete_house_user(request, house_user_pk):
    if request.user.is_anonymous or not request.user.role.building_access:
        return JsonResponse({'answer': 'Ви не маєте доступу до будинків'})
    try:
        house_user_to_delete = HouseUser.objects.get(pk=house_user_pk)
        house_user_to_delete.delete()
        return JsonResponse({'answer': 'success'})
    except (ValueError, AttributeError, HouseUser.DoesNotExist):
        return JsonResponse({'answer': 'failed'})


class FlatCreateView(PermissionCreateView):
    model = Flat
    template_name = 'administrator_panel/flat-create-update.html'
    form_class = FlatForm
    string_permission = 'flat_access'

    def get_context_data(self, **kwargs):
        context = super(FlatCreateView, self).get_context_data(**kwargs)
        houses = House.objects.prefetch_related('floor_set', 'section_set', 'flat_set').all()
        house_section = {}
        house_floor = {}
        for house in houses:
            house_section[house.pk] = []
            for section in house.section_set.all():
                house_section[house.pk].append([section.pk, section.name])
        for house in houses:
            house_floor[house.pk] = []
            for floor in house.floor_set.all():
                house_floor[house.pk].append([floor.pk, floor.name])
        context['owners'] = User.objects.filter(role__role='owner')
        context['tariffs'] = Tariff.objects.all()
        context['houses'] = houses
        context['house_section'] = house_section
        context['house_floor'] = house_floor
        context['personal_accounts'] = PersonalAccount.objects.filter(flat__isnull=True)
        if self.request.GET.get('prev_flat'):
            context['base_flat'] = Flat.objects.select_related('section',
                                                               'floor',
                                                               'house',
                                                               'owner').prefetch_related('house__floor_set',
                                                                                         'house__section_set').get(
                pk=self.request.GET.get('prev_flat'))
        return context

    def post(self, request, *args, **kwargs):
        self.object = None
        form = FlatForm(request.POST)
        if form.is_valid():
            return self.form_valid(form, personal_account_number=request.POST.get('personal-account'))
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)

    def form_valid(self, form, personal_account_number):
        created_flat = form.save()
        if personal_account_number:
            try:
                account = PersonalAccount.objects.get(number=personal_account_number)
                if not account.flat:
                    account.flat = created_flat
                    account.status = 'active'
                    account.save()
            except PersonalAccount.DoesNotExist:
                PersonalAccount.objects.create(number=personal_account_number,
                                               flat=created_flat,
                                               status='active')
        if self.request.POST.get('create-again') == 'create-new':
            return redirect(f'{reverse("flat-create")}?prev_flat={created_flat.pk}')
        return redirect('flat-detail', flat_pk=created_flat.pk)


class FlatDetailView(PermissionDetailView):
    model = Flat
    template_name = 'administrator_panel/flat-detail.html'
    pk_url_kwarg = 'flat_pk'
    string_permission = 'flat_access'

    def get_object(self, queryset=None):
        try:
            return Flat.objects.select_related('house', 'section', 'floor', 'owner', 'personalaccount').get(
                pk=self.kwargs['flat_pk'])
        except Flat.DoesNotExist:
            raise Http404()


class FlatListView(PermissionListView):
    model = Flat
    template_name = 'administrator_panel/flat-list.html'
    queryset = Flat.objects.select_related('house', 'section', 'floor', 'owner', 'personalaccount').all().order_by(
        '-id')
    string_permission = 'flat_access'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(FlatListView, self).get_context_data(object_list=object_list, **kwargs)
        personal_accounts = calculate_notoriety_and_receipt_sum(
            PersonalAccount.objects.select_related('flat').filter(flat__isnull=False, status='active')).get(
            'personal_accounts')
        context['object_list'] = context['object_list'] \
            .annotate(
            rest=Subquery(personal_accounts.filter(pk=OuterRef('personalaccount__id')).values('rest')[:1]),
            receipt_sum=Subquery(
                personal_accounts.filter(pk=OuterRef('personalaccount__id')).values('receipt_sum')[:1]),
            subtraction=Case(When(rest__isnull=True, then=Value(0.00)), default='rest', output_field=FloatField()) -
                        Case(When(receipt_sum__isnull=True, then=Value(0.00)), default='receipt_sum',
                             output_field=FloatField()))
        return context


class FlatUpdateView(PermissionUpdateView):
    model = Flat
    template_name = 'administrator_panel/flat-create-update.html'
    form_class = FlatForm
    pk_url_kwarg = 'flat_pk'
    string_permission = 'flat_access'

    def get_object(self, queryset=None):
        try:
            return Flat.objects.get(pk=self.kwargs['flat_pk'])
        except Flat.DoesNotExist:
            raise Http404()

    def get_context_data(self, **kwargs):
        flat = self.get_object()
        context = super(FlatUpdateView, self).get_context_data(**kwargs)
        houses = House.objects.prefetch_related('floor_set', 'section_set', 'flat_set').all()
        house_section = {}
        house_floor = {}
        for house in houses:
            house_section[house.pk] = []
            for section in house.section_set.all():
                house_section[house.pk].append([section.pk, section.name])
        for house in houses:
            house_floor[house.pk] = []
            for floor in house.floor_set.all():
                house_floor[house.pk].append([floor.pk, floor.name])
        try:
            context['selected_account'] = PersonalAccount.objects.get(flat=flat)
        except PersonalAccount.DoesNotExist:
            context['selected_account'] = None
        context['owners'] = User.objects.filter(role__role='owner')
        context['tariffs'] = Tariff.objects.all()
        context['houses'] = houses
        context['house_section'] = house_section
        context['house_floor'] = house_floor
        context['personal_accounts'] = PersonalAccount.objects.filter(flat__isnull=True)
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = FlatForm(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form, request.POST.get('personal-account'))
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)

    def form_valid(self, form, personal_account_number):
        saved_flat = form.save()
        try:
            personal_account = PersonalAccount.objects.get(flat=saved_flat)
            if personal_account.number != personal_account_number:
                personal_account.flat = None
                personal_account.save()
                if personal_account_number:
                    try:
                        new_personal_account = PersonalAccount.objects.get(number=personal_account_number)
                        new_personal_account.flat = saved_flat
                        new_personal_account.save()
                    except PersonalAccount.DoesNotExist:
                        PersonalAccount.objects.create(number=personal_account_number,
                                                       flat=saved_flat)
        except PersonalAccount.DoesNotExist:
            if personal_account_number:
                try:
                    personal_account = PersonalAccount.objects.get(number=personal_account_number)
                    personal_account.flat = saved_flat
                    personal_account.save()
                except PersonalAccount.DoesNotExist:
                    PersonalAccount.objects.create(number=personal_account_number,
                                                   flat=saved_flat)
        return redirect('flat-detail', flat_pk=saved_flat.pk)


def delete_flat(request, flat_pk):
    if request.user.is_anonymous or not request.user.role.flat_access:
        return JsonResponse({'answer': 'Ви не маєте доступу до квартир'})
    try:
        flat_to_delete = Flat.objects.get(pk=flat_pk)
        flat_to_delete.delete()
        return JsonResponse({'answer': 'success'})
    except (Flat.DoesNotExist, KeyError, AttributeError, ProtectedError):
        return JsonResponse({'answer': 'failed'})


def flat_number_is_unique(request):
    try:
        Flat.objects.get(number=request.GET.get('number'))
        return JsonResponse({'answer': 'failed'})
    except Flat.DoesNotExist:
        return JsonResponse({'answer': 'success'})


def personal_account_context(context, flats=Flat.objects.none()):
    """Separate function for forming context in both CreateView and UpdateView of PersonalAccount instances"""

    flats = flats
    house_section: dict[
        int, list[list[int, str]]] = {}  # used in frontend for dynamical changing of sections while house was changed
    section_flat: dict[
        int, list[list[int, str]]] = {}  # used in frontend for dynamical changing of flats while section was changed
    houses = []  # used for displaying houses in select tag in frontend

    for flat in flats:
        houses.append(flat.house)
        if not house_section.get(flat.house.id):
            house_section[flat.house.id] = [[flat.section.id, flat.section.name]]
        else:
            if [flat.section.id, flat.section.name] not in house_section[flat.house.id]:
                house_section[flat.house.id].append([flat.section.id, flat.section.name])

        if not section_flat.get(flat.section.id):
            section_flat[flat.section.id] = [[flat.id, flat.number]]
        else:
            section_flat[flat.section.id].append([flat.id, flat.number])

    context['flats'] = flats
    context['house_section'] = house_section
    context['section_flat'] = section_flat
    context['houses'] = set(houses)
    return context


class PersonalAccountCreateView(PermissionCreateView):
    model = PersonalAccount
    template_name = 'administrator_panel/personal_account-create-update.html'
    form_class = PersonalAccountForm
    string_permission = 'bill_access'

    def get_context_data(self, **kwargs):
        context = super(PersonalAccountCreateView, self).get_context_data(**kwargs)
        context = personal_account_context(context, Flat.objects.select_related('personalaccount', 'house', 'section',
                                                                                'owner').filter(
            personalaccount__isnull=True))
        context['create_new'] = {'create': 'true'}
        return context

    def post(self, request, *args, **kwargs):
        form = self.get_form_class()(request.POST)
        if form.is_valid():
            return self.form_valid(form)
        self.object = None
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)

    def form_valid(self, form):
        account = form.save()
        if account.flat:
            account.status = 'active'
            account.save()
        return redirect('personal-account-detail', account_pk=account.pk)


class PersonalAccountDetailView(PermissionDetailView):
    model = PersonalAccount
    template_name = 'administrator_panel/personal_account-detail.html'
    pk_url_kwarg = 'account_pk'
    string_permission = 'bill_access'

    def get_object(self, queryset=None):
        try:
            personal_account = PersonalAccount.objects.filter(pk=self.kwargs.get('account_pk'))
            return personal_account
        except PersonalAccount.DoesNotExist:
            raise Http404()

    def get_context_data(self, **kwargs):
        context = super(PersonalAccountDetailView, self).get_context_data(**kwargs)
        context.update(calculate_notoriety_and_receipt_sum(self.get_object()))
        context['object'] = context['personal_accounts'][0]
        return context


class PersonalAccountListView(PermissionListView):
    model = PersonalAccount
    template_name = 'administrator_panel/personal_account-list.html'
    string_permission = 'bill_access'

    def get_queryset(self):
        return PersonalAccount.objects \
            .select_related('flat', 'flat__section', 'flat__house', 'flat__owner') \
            .prefetch_related('notoriety_set', 'receipt_set', 'receipt_set__receiptservices') \
            .order_by('-id')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(PersonalAccountListView, self).get_context_data(object_list=object_list, **kwargs)
        context.update(count_all_totals(context['object_list']))

        return context


class PersonalAccountUpdateView(PermissionUpdateView):
    model = PersonalAccount
    template_name = 'administrator_panel/personal_account-create-update.html'
    pk_url_kwarg = 'account_pk'
    form_class = PersonalAccountForm
    string_permission = 'bill_access'

    def get_object(self, queryset=None):
        try:
            return PersonalAccount.objects.select_related('flat', 'flat__section', 'flat__house').prefetch_related(
                'flat__house__section_set').get(pk=self.kwargs['account_pk'])
        except PersonalAccount.DoesNotExist:
            raise Http404()

    def get_context_data(self, **kwargs):
        self.object = self.get_object()
        context = super(PersonalAccountUpdateView, self).get_context_data(**kwargs)
        context = personal_account_context(context, Flat.objects.select_related('personalaccount',
                                                                                'house',
                                                                                'section',
                                                                                'owner').filter(
            personalaccount__isnull=True))

        context['create_new'] = {'create': 'false'}

        # if flat related to current personal account exists
        try:
            # union Queryset<Flat> with flat related to current personal account for displaying it in select tag in frontend
            context['flats'] = context['flats'].union(
                Flat.objects.select_related('personalaccount', 'house', 'section', 'owner').filter(
                    pk=self.object.flat.pk))

            # add current flat to section's list of flats for displaying in frontend
            if context['section_flat'].get(self.object.flat.section.id):
                if [self.object.flat.id, self.object.flat.number] not in context['section_flat'][
                        self.object.flat.section.id]:
                    context['section_flat'][self.object.flat.section.id].append(
                        [self.object.flat.id, self.object.flat.number])
            else:
                context['section_flat'][self.object.flat.section.id] = [[self.object.flat.id, self.object.flat.number]]

            # add section of current personal account's flat to list of sections in separate house
            if self.object.flat.house not in context['houses']:
                context['house_section'][self.object.flat.house.pk] = [
                    [self.object.flat.section.pk, self.object.flat.section.name]]
                context['houses'].add(self.object.flat.house)
        except AttributeError:
            return context

        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form_class()(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)

    def form_valid(self, form):
        account_saved = form.save(commit=False)

        # if flat was deleted from account, it deletes flat from model instance
        if not form.cleaned_data.get('flat'):
            account_saved.flat = None

        # if account is bounded to flat, it cannot be inactive
        if form.cleaned_data.get('flat') and account_saved.status != 'active':
            account_saved.status = 'active'
        account_saved.save()
        return redirect('personal-account-detail', account_pk=account_saved.pk)


def delete_personal_account(request, account_pk):
    if request.user.is_anonymous or not request.user.role.bill_access:
        return JsonResponse({'answer': 'Ви не маєте доступу до особових рахунків'})
    try:
        account_to_delete = PersonalAccount.objects.prefetch_related('flat__evidence_set').get(pk=account_pk)
        if account_to_delete.flat.evidence_set.all().count() > 0:
            raise PersonalAccount.DoesNotExist()
        account_to_delete.delete()
        return JsonResponse({'answer': 'success'})
    except (PersonalAccount.DoesNotExist, KeyError, AttributeError, ProtectedError):
        return JsonResponse({'answer': 'failed'})


def personal_account_is_unique(request):
    try:
        PersonalAccount.objects.get(number=request.GET.get('number'), flat__isnull=False)
        return JsonResponse({'answer': 'failed'})
    except PersonalAccount.DoesNotExist:
        return JsonResponse({'answer': 'success'})


class PersonalAccountExcelView(View):

    def get(self, request, *args, **kwargs):
        if self.request.user.is_authenticated and self.request.user.role.bill_access:
            personal_accounts_dict = calculate_notoriety_and_receipt_sum(PersonalAccount.objects
                                                                         .select_related('flat', 'flat__section',
                                                                                         'flat__house', 'flat__owner') \
                                                                         .prefetch_related('notoriety_set',
                                                                                           'receipt_set',
                                                                                           'receipt_set__receiptservices')
                                                                         .all())
            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'].value = 'Особовий рахунок'
            ws['B1'].value = 'Статус'
            ws['C1'].value = 'Будинок'
            ws['D1'].value = 'Секція'
            ws['E1'].value = 'Квартира'
            ws['F1'].value = 'Власник'
            ws['G1'].value = 'Залишок'

            ws.column_dimensions['A'].width = 45
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 25
            ws.column_dimensions['D'].width = 20
            ws.column_dimensions['E'].width = 20
            ws.column_dimensions['F'].width = 30
            ws.column_dimensions['G'].width = 10

            for index, account in enumerate(personal_accounts_dict['personal_accounts']):
                ws[f'A{index + 2}'].value = account.number
                ws[f'B{index + 2}'].value = account.get_status_display()

                if account.flat:
                    ws[f'C{index + 2}'].value = account.flat.house.name
                    ws[f'D{index + 2}'].value = account.flat.section.name
                    ws[f'E{index + 2}'].value = account.flat.number
                    if account.flat.owner:
                        owner_string = account.flat.owner.surname + ' ' if account.flat.owner.surname else ''
                        owner_string += account.flat.owner.name + ' ' if account.flat.owner.name else ''
                        if account.flat.owner.father:
                            owner_string += account.flat.owner.father + ' '
                        ws[f'F{index + 2}'].value = owner_string

                ws[f'G{index + 2}'].value = account.subtraction if account.subtraction else 0

            from os import path, mkdir
            if not path.exists(f'{settings.MEDIA_ROOT}/accounts'):
                mkdir(f'{settings.MEDIA_ROOT}/accounts')

            file_name = f'accounts_{timezone.now().day}.{timezone.now().month}.{timezone.now().year}.xlsx'
            file_path = f'{settings.MEDIA_ROOT}/accounts/{file_name}'
            wb.save(file_path)
            return JsonResponse({'answer': 'success', 'file_path': f'{settings.MEDIA_URL}/accounts/{file_name}'})
        return JsonResponse({'answer': 'failed'})


class OwnerCreateView(PermissionCreateView):
    model = User
    template_name = 'administrator_panel/owner-create-update.html'
    form_class = OwnerForm
    string_permission = 'owner_access'

    def post(self, request, *args, **kwargs):
        form = self.get_form_class()(request.POST, request.FILES)
        if form.is_valid():
            return self.form_valid(form)
        self.object = None
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)

    def form_valid(self, form):
        User.objects.create_user(email=form.cleaned_data.get('email'),
                                 password=form.cleaned_data.get('password'),
                                 name=form.cleaned_data.get('name'),
                                 surname=form.cleaned_data.get('surname'),
                                 role=Role.objects.get(role__exact='owner'),
                                 birthday=form.cleaned_data.get('birthday'),
                                 logo=form.cleaned_data.get('logo'),
                                 father=form.cleaned_data.get('father'),
                                 phone=form.cleaned_data.get('phone'),
                                 viber=form.cleaned_data.get('viber'),
                                 owner_id=form.cleaned_data.get('owner_id'),
                                 telegram=form.cleaned_data.get('telegram'),
                                 status=form.cleaned_data.get('status'),
                                 notes=form.cleaned_data.get('notes'))
        return redirect('owners')


class OwnerUpdateView(PermissionUpdateView):
    model = User
    template_name = 'administrator_panel/owner-create-update.html'
    form_class = OwnerForm
    pk_url_kwarg = 'owner_pk'
    string_permission = 'owner_access'

    def get_object(self, queryset=None):
        try:
            return User.objects.select_related('role').get(pk=self.kwargs.get('owner_pk'), role__role='owner')
        except (User.DoesNotExist, KeyError, AttributeError):
            raise Http404()

    def post(self, request, *args, **kwargs):
        owner = self.get_object()
        form = OwnerForm(request.POST, request.FILES, instance=owner)
        if form.is_valid():
            return self.form_valid(form, owner)
        self.object = self.get_object()
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)

    def form_valid(self, form, owner):
        old_password = owner.password
        owner_saved = form.save()
        if form.cleaned_data.get('password'):
            owner_saved.set_password(form.cleaned_data.get('password'))
            owner_saved.save()
        else:
            owner_saved.password = old_password
            owner_saved.save()
        return redirect('owners')


class OwnerListView(PermissionListView):
    model = User
    queryset = User.objects.prefetch_related('flat_set', 'flat_set__house').filter(role__role='owner').order_by('-id')
    template_name = 'administrator_panel/owner-list.html'
    string_permission = 'owner_access'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=object_list, **kwargs)
        personal_accounts = calculate_notoriety_and_receipt_sum(PersonalAccount.objects
                                                                .select_related('flat', 'flat__section', 'flat__house',
                                                                                'flat__owner')
                                                                .prefetch_related('notoriety_set', 'receipt_set',
                                                                                  'receipt_set__receiptservices')
                                                                .filter(flat__owner__isnull=False))['personal_accounts']

        # dictionary with debt information about each owner
        context['debts'] = {}
        for owner in context['object_list']:
            context['debts'][owner.pk] = 'немає даних'
            for account in personal_accounts:
                if account.flat.owner.pk == owner.pk and account.subtraction < 0:
                    context['debts'][owner.pk] = 'Є'
                elif account.flat.owner.pk == owner.pk and account.subtraction >= 0:
                    context['debts'][owner.pk] = 'Немає'
        return context


class OwnerDetailView(PermissionDetailView):
    model = User
    template_name = 'administrator_panel/owner-detail.html'
    pk_url_kwarg = 'owner_pk'
    string_permission = 'owner_access'

    def get_object(self, queryset=None):
        try:
            return User.objects.get(pk=self.kwargs['owner_pk'])
        except User.DoesNotExist:
            raise Http404()


def delete_owner(request, owner_pk):
    if request.user.is_anonymous or not request.user.role.owner_access:
        return JsonResponse({'answer': 'Ви не маєте доступу до власників квартир'})
    try:
        user = User.objects.prefetch_related('flat_set__receipt_set', 'flat_set__personalaccount').get(pk=owner_pk)
        if not user.flat_set.all().exists():
            user.delete()
            return JsonResponse({'answer': 'success'})
        return JsonResponse({'answer': 'failed'})
    except User.DoesNotExist:
        return JsonResponse({'answer': 'failed'})


class SendInvitationByEmailView(TemplateResponseMixin, FormMixin, PermissionView):
    string_permission = 'owner_access'
    template_name = 'administrator_panel/send-invitation.html'
    form_class = SendInvitationForm
    name = 'надсилання email-листів'

    def get(self, request, *args, **kwargs):
        context = self.get_context_data()
        return self.render_to_response(context)

    def post(self, request, *args, **kwargs):
        form = self.get_form_class()(request.POST)
        if form.is_valid():
            return self.form_valid(form)
        return self.form_invalid(form)

    def form_valid(self, form):
        email_to_send = form.cleaned_data.get('email')
        subject = 'Запрошення в Мой Дом 24'
        main_text = 'Вас запрошують приєднатися до системи "Мой Дом 24".\n\nДля подальшої інформації зв`яжіться із адміністратором.'
        email = EmailMessage(subject, main_text, to=[email_to_send])
        if email.send(fail_silently=True):
            messages.success(self.request, 'Запрошення успішно надіслано.')
        else:
            messages.error(self.request, 'Повідомлення не було відправленим. Перевірте правильність вводу.')
        return redirect('send-owner-invitation')

    def form_invalid(self, form):
        messages.error(self.request, 'Щось пішло не так. Перевірте правильність вводу.')
        return redirect('send-owner-invitation')


class EvidenceCreateView(PermissionCreateView):
    model = Evidence
    template_name = 'administrator_panel/evidence-create-update.html'
    form_class = EvidenceForm
    string_permission = 'counter_access'

    def get_context_data(self, **kwargs):
        context = super(EvidenceCreateView, self).get_context_data(**kwargs)
        context = personal_account_context(context, Flat.objects.prefetch_related('house', 'section').filter(
            personalaccount__isnull=False))
        context['services'] = Service.objects.select_related('measurement_unit').filter(show_in_counters=True)
        context['create_new_number'] = {'create': 'true'}
        if self.request.GET.get('base_evidence'):
            context['base_evidence'] = Evidence.objects.select_related('flat', 'flat__house', 'flat__section',
                                                                       'service').prefetch_related(
                'flat__house__section_set', 'flat__section__flat_set').get(pk=self.request.GET.get('base_evidence'))
        return context

    def post(self, request, *args, **kwargs):
        form = self.get_form_class()(request.POST)
        if form.is_valid():
            return self.form_valid(form)
        self.object = None
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)

    def form_valid(self, form):
        evidence_saved = form.save()
        if self.request.POST.get('create-again'):
            return redirect(f"{reverse('evidence-create')}?base_evidence={evidence_saved.pk}")
        return redirect('evidence-detail', evidence_pk=evidence_saved.pk)


class EvidenceDetailView(PermissionDetailView):
    model = Evidence
    template_name = 'administrator_panel/evidence-detail.html'
    pk_url_kwarg = 'evidence_pk'
    string_permission = 'counter_access'


class GroupedEvidenceListView(PermissionListView):
    model = Evidence
    template_name = 'administrator_panel/evidence-grouped-list.html'
    queryset = Evidence.objects.select_related('service',
                                               'flat',
                                               'flat__house',
                                               'flat__section',
                                               'service__measurement_unit').all().annotate(
        distinct_name=Concat('flat', 'service')).distinct('distinct_name')
    string_permission = 'counter_access'


class ServiceEvidenceListView(PermissionListView):
    model = Evidence
    template_name = 'administrator_panel/evidence-counter-list.html'
    string_permission = 'counter_access'

    def get_queryset(self):
        flat_pk = self.request.GET.get('flat')
        service_pk = self.request.GET.get('service')
        try:
            return Evidence.objects.select_related('flat', 'service', 'flat__house', 'flat__section',
                                                   'service__measurement_unit').filter(
                flat=Flat.objects.get(pk=flat_pk), service=Service.objects.get(pk=service_pk)).order_by('-id')
        except (Evidence.DoesNotExist, Flat.DoesNotExist, Service.DoesNotExist, KeyError, AttributeError):
            raise Http404()


class EvidenceUpdateView(PermissionUpdateView):
    model = Evidence
    template_name = 'administrator_panel/evidence-create-update.html'
    form_class = EvidenceForm
    pk_url_kwarg = 'evidence_pk'
    string_permission = 'counter_access'

    def get_object(self, queryset=None):
        try:
            return Evidence.objects.select_related('flat',
                                                   'service',
                                                   'flat__house',
                                                   'flat__section',
                                                   'flat__owner').get(pk=self.kwargs.get('evidence_pk'))
        except (Evidence.DoesNotExist, AttributeError, KeyError):
            raise Http404()

    def get_context_data(self, **kwargs):
        context = super(EvidenceUpdateView, self).get_context_data(**kwargs)
        context = personal_account_context(context, Flat.objects.prefetch_related('house', 'section').filter(
            personalaccount__isnull=False))
        context['services'] = Service.objects.select_related('measurement_unit').filter(show_in_counters=True)
        context['create_new_number'] = {'create': 'false'}
        return context

    def post(self, request, *args, **kwargs):
        form = self.get_form_class()(request.POST, instance=self.get_object())
        if form.is_valid():
            return self.form_valid(form)
        self.object = self.get_object()
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)

    def form_valid(self, form):
        evidence_saved = form.save()
        return redirect('evidence-detail', evidence_pk=evidence_saved.pk)


def delete_evidence(request, evidence_pk):
    if request.user.is_anonymous or not request.user.role.counter_access:
        return JsonResponse({'answer': 'Ви не маєте доступу до показників лічильників'})
    try:
        evidence_to_delete = Evidence.objects.get(pk=evidence_pk)
        evidence_to_delete.delete()
        return JsonResponse({'answer': 'success'})
    except Evidence.DoesNotExist:
        return JsonResponse({'answer': 'failed'})


class ReceiptCreateView(PermissionCreateView):
    model = Receipt
    template_name = 'administrator_panel/receipt-create-update.html'
    form_class = ReceiptForm
    string_permission = 'receipt_access'

    def get_context_data(self, **kwargs):
        context = super(ReceiptCreateView, self).get_context_data(**kwargs)
        context = personal_account_context(context,
                                           flats=Flat.objects.select_related('house', 'section', 'personalaccount',
                                                                             'tariff', 'owner',
                                                                             'section__house').filter(
                                               personalaccount__isnull=False, personalaccount__status='active'))

        # adding sections of each house, whose flats are with personal account in order to display it on firstly loaded page
        if self.request.GET.get('base_receipt') or self.request.GET.get('flat_id'):
            context['sections_in_houses'] = set()
            for flat in context['flats']:
                context['sections_in_houses'].add(flat.section)

            if self.request.GET.get('flat_id'):
                # trying to get flat, on which new receipt will be based
                try:
                    context['base_flat'] = Flat.objects.select_related('personalaccount',
                                                                       'house',
                                                                       'section',
                                                                       'tariff',
                                                                       'owner').get(pk=self.request.GET.get('flat_id'))
                except Flat.DoesNotExist:
                    raise Http404()

            if self.request.GET.get('base_receipt'):
                # getting receipt basing on which new receipt will be built
                try:
                    context['base_receipt'] = Receipt.objects.select_related('account',
                                                                             'account__flat',
                                                                             'account__flat__house',
                                                                             'account__flat__section',
                                                                             'account__flat__owner').get(
                        pk=self.request.GET.get('base_receipt'))
                    context['base_receipt_services'] = ReceiptService.objects.select_related('service',
                                                                                             'receipt',
                                                                                             'service__measurement_unit').filter(
                        receipt=context['base_receipt'])
                except Receipt.DoesNotExist:
                    raise Http404()

        context['personal_accounts'] = PersonalAccount.objects.select_related('flat', 'flat__owner').filter(
            flat__isnull=False)
        context['receipt_service_formset'] = receipt_service_formset(queryset=ReceiptService.objects.none())
        context['tariffs'] = Tariff.objects.prefetch_related('tariffservice_set',
                                                             'tariffservice_set__service',
                                                             'tariffservice_set__service__measurement_unit').all()
        context['services'] = Service.objects.select_related('measurement_unit').all()
        context['flat_personal_account'] = {}
        context['create_new_number'] = {'create': 'true'}
        context['flat_tariff'] = {}

        # dictionary where flat.pk is key and flat's personal account is value. If no account - 'none'
        for flat in context['flats']:
            context['flat_personal_account'][flat.id] = flat.personalaccount.number
            if flat.tariff:
                context['flat_tariff'][flat.id] = flat.tariff.id
            else:
                context['flat_tariff'][flat.id] = 'none'

        return context

    def post(self, request, *args, **kwargs):
        form = self.get_form_class()(self.request.POST)
        receipt_formset = receipt_service_formset(request.POST, queryset=ReceiptService.objects.none())
        if form.is_valid() and receipt_formset.is_valid():
            return self.form_valid(form, receipt_formset=receipt_formset)
        self.object = None
        context = self.get_context_data()
        context['form'] = form
        context['receipt_service_formset'] = receipt_formset
        context['create_new_number'] = {'create': 'false'}
        return self.render_to_response(context)

    def form_valid(self, form, receipt_formset):
        receipt_saved = form.save()
        for receipt_service_form in receipt_formset.forms:
            if receipt_service_form.cleaned_data.get('service') and receipt_service_form.cleaned_data.get('amount') \
                    and receipt_service_form.cleaned_data.get('price') and receipt_service_form.cleaned_data.get('total_price'):
                form_saved = receipt_service_form.save(commit=False)
                form_saved.receipt = receipt_saved
                form_saved.save()
        return redirect('receipt-detail', receipt_pk=receipt_saved.pk)


class ReceiptUpdateView(PermissionUpdateView):
    model = Receipt
    template_name = 'administrator_panel/receipt-create-update.html'
    pk_url_kwarg = 'receipt_id'
    form_class = ReceiptForm
    string_permission = 'receipt_access'

    def get_object(self, queryset=None):
        try:
            return Receipt.objects.select_related('account',
                                                  'account__flat',
                                                  'account__flat__house',
                                                  'account__flat__section').prefetch_related('receiptservices').get(
                pk=self.kwargs.get('receipt_pk'))
        except Receipt.DoesNotExist:
            raise Http404()

    def get_context_data(self, **kwargs):
        self.object = self.get_object()
        context = super(ReceiptUpdateView, self).get_context_data(**kwargs)
        context = personal_account_context(context,
                                           flats=Flat.objects.select_related('house', 'section', 'personalaccount',
                                                                             'tariff', 'owner',
                                                                             'section__house').filter(
                                               personalaccount__isnull=False, personalaccount__status='active'))

        # adding sections of each house, whose flats are with personal account in order to display it on firstly loaded page
        context['sections_in_houses'] = set()
        for flat in context['flats']:
            context['sections_in_houses'].add(flat.section)

        context['personal_accounts'] = PersonalAccount.objects.select_related('flat', 'flat__owner').filter(
            flat__isnull=False)
        context['receipt_service_formset'] = receipt_service_formset(
            queryset=ReceiptService.objects.select_related('service', 'service__measurement_unit').filter(
                receipt=self.object))
        context['tariffs'] = Tariff.objects.prefetch_related('tariffservice_set',
                                                             'tariffservice_set__service',
                                                             'tariffservice_set__service__measurement_unit').all()
        context['services'] = Service.objects.select_related('measurement_unit').all()
        context['flat_personal_account'] = {}
        context['create_new_number'] = {'create': 'false'}
        context['flat_tariff'] = {}

        # dictionary where flat.pk is key and flat's personal account is value. If no account - 'none'
        for flat in context['flats']:
            context['flat_personal_account'][flat.id] = flat.personalaccount.number
            if flat.tariff:
                context['flat_tariff'][flat.id] = flat.tariff.id
            else:
                context['flat_tariff'][flat.id] = 'none'

        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form_class()(request.POST, instance=self.object)
        receipt_formset = receipt_service_formset(request.POST,
                                                  queryset=ReceiptService.objects.filter(receipt=self.object))
        if form.is_valid() and receipt_formset.is_valid():
            return self.form_valid(form, receipt_formset=receipt_formset)
        context = self.get_context_data()
        context['form'] = form
        context['receipt_service_formset'] = receipt_formset
        return self.render_to_response(context)

    def form_valid(self, form, receipt_formset):
        receipt_saved = form.save()
        for receipt_service_form in receipt_formset.forms:
            if receipt_service_form.cleaned_data.get('service') and receipt_service_form.cleaned_data.get('amount') \
                    and receipt_service_form.cleaned_data.get('price') and receipt_service_form.cleaned_data.get('total_price'):
                form_saved = receipt_service_form.save(commit=False)
                form_saved.receipt = receipt_saved
                form_saved.save()
        return redirect('receipt-detail', receipt_pk=receipt_saved.pk)


class ReceiptDetailView(PermissionDetailView):
    model = Receipt
    template_name = 'administrator_panel/receipt-detail.html'
    pk_url_kwarg = 'receipt_pk'
    string_permission = 'receipt_access'

    def get_object(self, queryset=None):
        try:
            return Receipt.objects.select_related('account',
                                                  'account__flat',
                                                  'account__flat__owner',
                                                  'account__flat__house',
                                                  'account__flat__section') \
                .prefetch_related('receiptservices',
                                  'receiptservices__service',
                                  'receiptservices__service__measurement_unit') \
                .annotate(summ=Sum('receiptservices__total_price')) \
                .get(pk=self.kwargs.get('receipt_pk'))
        except (Receipt.DoesNotExist, KeyError, AttributeError):
            raise Http404()


class ReceiptListView(PermissionListView):
    model = Receipt
    template_name = 'administrator_panel/receipt-list.html'
    string_permission = 'receipt_access'

    def get_queryset(self):
        return Receipt.objects \
            .select_related('account', 'account__flat', 'account__flat__house', 'account__flat__section',
                            'account__flat__owner') \
            .prefetch_related('receiptservices').all().order_by('-id') \
            .annotate(summ=Sum('receiptservices__total_price'))

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(ReceiptListView, self).get_context_data(object_list=object_list, **kwargs)
        context.update(count_all_totals(PersonalAccount.objects.all()))

        return context


class EvidenceResponse(MultipleObjectMixin, View):
    model = Evidence

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous or not request.user.role.counter_access:
            return JsonResponse({'answer': 'Ви не маєте доступу до показань лічильників'})
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        try:
            evidences = Evidence.objects.select_related('flat__personalaccount',
                                                        'flat',
                                                        'flat__house',
                                                        'flat__section',
                                                        'service',
                                                        'service_id',
                                                        'service__measurement_unit') \
                .filter(flat__personalaccount__number=self.request.GET.get('account_number')) \
                .values('number',
                        'status',
                        'date_from',
                        'flat__house__name',
                        'flat__section__name',
                        'flat__number',
                        'service__name',
                        'counter_evidence',
                        'service__measurement_unit__name',
                        'service_id')
            return list(evidences)
        except (KeyError, AttributeError):
            return None

    def get_context_data(self, **kwargs):
        evidences = self.get_queryset()
        for evidence in evidences:
            evidence['date_from_formatted'] = evidence['date_from'].strftime('%d.%m.%Y')
            evidence['date_from_month'] = evidence['date_from'].strftime('%B').title()
        return JsonResponse({'evidences': evidences})

    def get(self, request):
        context = self.get_context_data()
        return context


def receipt_service_delete(request, receipt_service_delete_pk):
    if request.user.is_anonymous or not request.user.role.receipt_access:
        return JsonResponse({'answer': 'Ви не маєте доступу до квитанцій'})
    try:
        receipt_service_to_delete = ReceiptService.objects.get(pk=receipt_service_delete_pk)
        receipt_service_to_delete.delete()
        return JsonResponse({'answer': 'success'})
    except (ReceiptService.DoesNotExist, KeyError, AttributeError):
        return JsonResponse({'answer': 'failed'})


def delete_receipt(request):
    if request.user.is_anonymous or not request.user.role.receipt_access:
        return JsonResponse({'answer': 'Ви не маєте доступу до квитанцій'})
    if request.is_ajax():
        try:
            receipt_to_delete = Receipt.objects.get(pk=request.GET.get('receipt_id'))
            receipt_to_delete.delete()
            return JsonResponse({'answer': 'success'})
        except (Receipt.DoesNotExist, AttributeError, KeyError):
            return JsonResponse({'answer': 'failed'})

    if request.POST.get('receipt_id'):
        try:
            receipt_to_delete = Receipt.objects.get(pk=request.POST.get('receipt_id'))
            receipt_to_delete.delete()
        finally:
            return redirect('receipts')

    if request.POST.get('receipts_to_delete'):
        receipts_pk = request.POST.get('receipts_to_delete').split(',')[:-1]
        try:
            receipts = Receipt.objects.filter(pk__in=receipts_pk)
            receipts.delete()
        finally:
            return redirect('receipts')
    return redirect('receipts')


class NotorietyCreateView(PermissionCreateView):
    model = Notoriety
    template_name = 'administrator_panel/notoriety-create-update.html'
    form_class = NotorietyForm
    string_permission = 'checkout_access'

    def get_context_data(self, **kwargs):
        context = super(NotorietyCreateView, self).get_context_data(**kwargs)
        if kwargs.get('type'):
            context['type'] = kwargs.get('type')
        else:
            context['type'] = self.request.GET.get('type') if self.request.GET.get(
                'type') else 'outcome'  # if type is not defined it is set 'outcome'

        # create notoriety based on already existing notoriety
        if self.request.GET.get('base_notoriety'):
            try:
                context['base_notoriety'] = Notoriety.objects \
                    .select_related('manager', 'manager__role', 'account', 'account__flat__owner', 'article') \
                    .get(pk=self.request.GET.get('base_notoriety'))
            except Notoriety.DoesNotExist:
                raise Http404()

        # create notoriety with indicated personal account
        if self.request.GET.get('account'):
            try:
                context['base_account'] = PersonalAccount.objects \
                    .select_related('flat', 'flat__owner') \
                    .get(pk=self.request.GET.get('account'))
                context['type'] = 'income'
            except (PersonalAccount.DoesNotExist, KeyError, AttributeError):
                raise Http404()

        context['personal_accounts'] = PersonalAccount.objects.select_related('flat', 'flat__owner') \
            .filter(flat__isnull=False, status='active')
        context['owners'] = set([account.flat.owner for account in context['personal_accounts'] if account.flat.owner])
        context['articles'] = ArticlePayment.objects.filter(type__exact=context['type'])
        context['managers'] = User.objects.select_related('role').filter(
            role__role__in=['director', 'manager', 'accountant'])

        context['create_new'] = {'create': 'true'}
        return context

    def post(self, request, *args, **kwargs):
        form = self.get_form_class()(request.POST)
        if form.is_valid():
            return self.form_valid(form)
        self.object = None
        context = self.get_context_data(type=self.request.POST.get('type'))
        context['form'] = form
        context['create_new'] = {'create': 'false'}
        return self.render_to_response(context)

    def form_valid(self, form):
        notoriety_saved = form.save()
        return redirect('notoriety-detail', notoriety_pk=notoriety_saved.pk)


class NotorietyDetailView(PermissionDetailView):
    model = Notoriety
    template_name = 'administrator_panel/notoriety-detail.html'
    pk_url_kwarg = 'notoriety_pk'
    string_permission = 'checkout_access'

    def get_object(self, queryset=None):
        try:
            return Notoriety.objects.select_related('account',
                                                    'account__flat',
                                                    'account__flat__owner',
                                                    'manager',
                                                    'manager__role',
                                                    'article').get(pk=self.kwargs.get('notoriety_pk'))
        except (Notoriety.DoesNotExist, KeyError, AttributeError):
            return None


class NotorietyTemplateDownload(SingleObjectMixin, View):
    model = Notoriety

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous or not request.user.role.checkout_access:
            return JsonResponse({'answer': 'Ви не маєте доступу до каси'})
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            notoriety = Notoriety.objects \
                .select_related('account', 'account__flat__owner', 'article') \
                .get(pk=request.POST.get('notoriety'))
        except (Notoriety.DoesNotExist, AttributeError):
            return JsonResponse({'answer': 'failed'})

        wb = openpyxl.Workbook()
        ws = wb.active

        ws['A1'].value = 'Платіж'
        ws['A2'].value = 'Дата'
        ws['A3'].value = 'Власник квартири'
        ws['A4'].value = 'Особовий рахунок'
        ws['A5'].value = 'Статус'
        ws['A6'].value = 'Стаття'
        ws['A7'].value = 'Сума'
        ws['A8'].value = 'Валюта'
        ws['A9'].value = 'Коментар'
        ws['A10'].value = 'Менеджер'

        ws['B1'].value = f'#{notoriety.number}'
        ws['B2'].value = f'{notoriety.created_at.day}.{notoriety.created_at.month}.{notoriety.created_at.year}'
        if notoriety.type == 'income':
            if notoriety.account.flat.owner:
                owner_string = notoriety.account.flat.owner.surname + ' ' if notoriety.account.flat.owner.surname else ''
                owner_string += notoriety.account.flat.owner.name + ' ' if notoriety.account.flat.owner.name else ''
                owner_string += notoriety.account.flat.owner.father if notoriety.account.flat.owner.father else ''
                ws['B3'].value = owner_string
            ws['B4'].value = notoriety.account.number if notoriety.account else ''
        ws['B5'].value = 'Проведена' if notoriety.is_completed else 'Не проведена'
        ws['B6'].value = notoriety.article.name
        ws['B7'].value = notoriety.sum if notoriety.type == 'income' else -abs(notoriety.sum)
        ws['B8'].value = 'UAH'
        ws['B9'].value = notoriety.comment
        ws['B10'].value = f'{notoriety.manager.surname} {notoriety.manager.name} {notoriety.manager.father}'

        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 35
        ws['B7'].alignment = Alignment(horizontal="left")

        from os import path, mkdir
        if not path.exists(f'{settings.MEDIA_ROOT}/notorieties'):
            mkdir(f'{settings.MEDIA_ROOT}/notorieties')

        file_name = f'notoriety_{notoriety.number}_{notoriety.created_at.day}.{notoriety.created_at.month}.{notoriety.created_at.year}.xlsx'
        file_path = f'{settings.MEDIA_ROOT}/notorieties/{file_name}'
        wb.save(file_path)
        return JsonResponse({'answer': 'success', 'file_path': f'{settings.MEDIA_URL}/notorieties/{file_name}'})


class NotorietyUpdateView(PermissionUpdateView):
    model = Notoriety
    template_name = 'administrator_panel/notoriety-create-update.html'
    pk_url_kwarg = 'notoriety_pk'
    form_class = NotorietyForm
    string_permission = 'checkout_access'

    def get_object(self, queryset=None):
        try:
            return Notoriety.objects \
                .select_related('account', 'account__flat', 'account__flat__owner', 'manager', 'manager__role',
                                'article') \
                .get(pk=self.kwargs.get('notoriety_pk'))
        except (Notoriety.DoesNotExist, KeyError, AttributeError):
            return None

    def get_context_data(self, **kwargs):
        context = super(NotorietyUpdateView, self).get_context_data(**kwargs)
        self.object = self.get_object()
        context['type'] = self.object.type
        context['create_new'] = {'create': 'false'}
        context['personal_accounts'] = PersonalAccount.objects.select_related('flat', 'flat__owner') \
            .filter(flat__isnull=False, status='active', flat__owner__isnull=False)
        context['owners'] = set([account.flat.owner for account in context['personal_accounts']])
        context['articles'] = ArticlePayment.objects.filter(type__exact=context['type'])
        context['managers'] = User.objects.select_related('role').filter(
            role__role__in=['director', 'manager', 'accountant'])
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form_class()(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)

    def form_valid(self, form):
        notoriety_saved = form.save()
        return redirect('notoriety-detail', notoriety_pk=notoriety_saved.pk)


class NotorietyListView(PermissionListView):
    model = Notoriety
    template_name = 'administrator_panel/notoriety-list.html'
    queryset = Notoriety.objects.select_related('article', 'account', 'account__flat__owner').filter(is_completed=True).order_by(
        '-created_at')
    string_permission = 'checkout_access'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super(NotorietyListView, self).get_context_data(object_list=object_list, **kwargs)
        context.update(count_all_totals(PersonalAccount.objects.all()))

        return context


class NotorietyListTemplateDownload(View):

    def get(self, request, *args, **kwargs):
        if self.request.user.is_authenticated and self.request.user.role.checkout_access:

            notorieties = Notoriety.objects \
                .select_related('article', 'account', 'account__flat', 'account__flat__owner') \
                .all()

            wb = openpyxl.Workbook()
            ws = wb.active
            ws['A1'].value = '#'
            ws['B1'].value = 'Дата'
            ws['C1'].value = 'Прихід/розхід'
            ws['D1'].value = 'Статус'
            ws['E1'].value = 'Стаття'
            ws['F1'].value = 'Сума'
            ws['G1'].value = 'Валюта'
            ws['H1'].value = 'Власник квартири'
            ws['I1'].value = 'Особовий рахунок'

            for index, notoriety in enumerate(notorieties):
                ws[f'A{index + 2}'].value = notoriety.number
                ws[
                    f'B{index + 2}'].value = f'{notoriety.created_at.day}.{notoriety.created_at.month}.{notoriety.created_at.year}'
                ws[f'C{index + 2}'].value = notoriety.get_type_display()
                ws[f'D{index + 2}'].value = 'Проведена' if notoriety.is_completed else 'Не проведена'
                ws[f'E{index + 2}'].value = notoriety.article.name
                ws[f'F{index + 2}'].value = notoriety.sum if notoriety.type == 'income' else -abs(notoriety.sum)
                ws[f'G{index + 2}'].value = 'UAH'
                if notoriety.type == 'income' and notoriety.account.flat.owner:
                    owner_string = notoriety.account.flat.owner.surname + ' ' if notoriety.account.flat.owner.surname else ''
                    owner_string += notoriety.account.flat.owner.name + ' ' if notoriety.account.flat.owner.name else ''
                    if notoriety.account.flat.owner.father:
                        owner_string += notoriety.account.flat.owner.father
                    ws[f'H{index + 2}'].value = owner_string
                ws[f'I{index + 2}'].value = notoriety.account.number if notoriety.type == 'income' else None

            ws.column_dimensions['A'].width = 45
            ws.column_dimensions['B'].width = 15
            ws.column_dimensions['C'].width = 15
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 35
            ws.column_dimensions['F'].width = 15
            ws.column_dimensions['G'].width = 10
            ws.column_dimensions['H'].width = 30
            ws.column_dimensions['I'].width = 30

            from os import path, mkdir
            if not path.exists(f'{settings.MEDIA_ROOT}/notorieties'):
                mkdir(f'{settings.MEDIA_ROOT}/notorieties')

            file_name = f'notorieties_{timezone.now().day}.{timezone.now().month}.{timezone.now().year}.xlsx'
            file_path = f'{settings.MEDIA_ROOT}/notorieties/{file_name}'
            wb.save(file_path)
            return JsonResponse({'answer': 'success', 'file_path': f'{settings.MEDIA_URL}/notorieties/{file_name}'})
        return JsonResponse({'answer': 'failed'})


class NotorietyDeleteView(SingleObjectMixin, View):
    model = Notoriety
    pk_url_kwarg = 'notoriety_pk'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous or not request.user.role.checkout_access:
            return JsonResponse({'answer': 'Ви не маєте доступу до каси'})
        return super().dispatch(request, *args, **kwargs)

    def get_queryset(self):
        return Notoriety.objects.all()

    def get_object(self, queryset=None):
        try:
            return queryset.get(pk=self.kwargs.get('notoriety_pk'))
        except (Notoriety.DoesNotExist, KeyError, AttributeError):
            return None

    def get(self, request, *args, **kwargs):
        self.object = self.get_object(self.get_queryset())
        try:
            self.object.delete()
            if self.request.is_ajax():
                return JsonResponse({'answer': 'success'})
            return redirect('notorieties')
        except (Notoriety.DoesNotExist, KeyError, AttributeError):
            if self.request.is_ajax():
                return JsonResponse({'answer': 'failed'})


class ReceiptTemplateCreateView(PermissionCreateView):
    model = Template
    template_name = 'administrator_panel/template-create.html'
    form_class = ReceiptTemplateForm
    string_permission = 'receipt_access'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['object_list'] = Template.objects.all().order_by('id')
        return context

    def post(self, request, *args, **kwargs):
        form = self.get_form_class()(request.POST, request.FILES)
        if form.is_valid():
            form.save()
            return redirect('template-create')
        return redirect('template-create')


class TemplateDefault(SingleObjectMixin, View):
    model = Template
    pk_url_kwarg = 'template_pk'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous or not request.user.role.receipt_access:
            return JsonResponse({'answer': 'Ви не маєте доступу до квитанцій'})
        return super().dispatch(request, *args, **kwargs)

    def get_object(self, queryset=None):
        try:
            return Template.objects.get(pk=self.kwargs.get('template_pk'))
        except Template.DoesNotExist:
            return None

    def get(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj:
            try:
                another_default = Template.objects.get(is_default=True)
                another_default.is_default = False
                another_default.save()
            except Template.DoesNotExist:
                pass
            obj.is_default = True
            obj.save()
            return redirect('template-create')
        return redirect('template-create')


class TemplateDeleteView(SingleObjectMixin, View):
    model = Template
    pk_url_kwarg = 'template_pk'

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous or not request.user.role.receipt_access:
            return JsonResponse({'answer': 'Ви не маєте доступу до квитанцій'})
        return super().dispatch(request, *args, **kwargs)

    def get_object(self, queryset=None):
        try:
            return Template.objects.get(pk=self.kwargs.get('template_pk'))
        except Template.DoesNotExist:
            return None

    def get(self, request, *args, **kwargs):
        obj = self.get_object()
        if obj:
            if not obj.is_default:
                obj.delete()
                return JsonResponse({'answer': 'success'})
            return JsonResponse({'answer': 'template is default'})
        return JsonResponse({'answer': 'failed'})


class TemplateChooseView(PermissionListView):
    model = Template
    queryset = Template.objects.all()
    template_name = 'administrator_panel/template-download.html'
    string_permission = 'receipt_access'

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=object_list)
        context['receipt'] = self.request.GET.get('receipt')
        return context


class BuildReceiptFileView(SingleObjectMixin, View):
    model = Receipt

    def dispatch(self, request, *args, **kwargs):
        if request.user.is_anonymous or (not request.user.role.receipt_access and request.user.role.role != 'owner'):
            return JsonResponse({'answer': 'Ви не маєте доступу'})
        return super().dispatch(request, *args, **kwargs)

    def post(self, request, *args, **kwargs):
        try:
            if request.GET.get('template'):
                template = Template.objects.get(pk=request.GET.get('template')).file
            else:
                template = Template.objects.get(is_default=True).file

            receipt = Receipt.objects.select_related('account',
                                                     'account__flat',
                                                     'account__flat__owner') \
                .prefetch_related('receiptservices',
                                  'receiptservices__service',
                                  'receiptservices__service__measurement_unit',
                                  'account__notoriety_set') \
                .filter(pk=request.GET.get('receipt'))
            if receipt.count() == 0:
                raise Receipt.DoesNotExist()

            # get account for calculating the total sum on account
            personal_account = PersonalAccount.objects \
                .select_related('flat', 'flat__section', 'flat__house', 'flat__owner') \
                .prefetch_related('notoriety_set', 'receipt_set', 'receipt_set__receiptservices') \
                .filter(pk=receipt[0].account.pk)
            if personal_account.count() == 0:
                raise PersonalAccount.DoesNotExist()

            personal_account = calculate_notoriety_and_receipt_sum(personal_account)['personal_accounts']
            receipt = receipt.annotate(receipt_sum=Sum('receiptservices__total_price'))

            receipt = receipt[0]
            payment_requisites = PaymentRequisite.objects.first()

            wb = openpyxl.Workbook()
            base_template = openpyxl.load_workbook(filename=template.path)
            base_sheet = base_template.active

            if receipt.account.flat.owner:
                name_surname_father = f'{receipt.account.flat.owner.surname} {receipt.account.flat.owner.name}'
                if receipt.account.flat.owner.father:
                    name_surname_father += f' {receipt.account.flat.owner.father}'
            else:
                name_surname_father = 'N/A'

            reserved_words = {
                '%payCompany%': payment_requisites.name if payment_requisites else 'N/A',
                '%accountNumber%': receipt.account.number,
                '%invoiceNumber%': receipt.number,
                '%invoiceDate%': f'{receipt.date_from.day}.{receipt.date_from.month}.{receipt.date_from.year}',
                '%invoiceAddress%': name_surname_father,
                '%accountBalance%': personal_account[0].subtraction,
                '%total%': receipt.receipt_sum if receipt.receipt_sum else 0.00,
                '%invoiceMonth%': f'{receipt.date_from.strftime("%B")} {receipt.date_from.year}',
                '%totalDebt%': abs(personal_account[0].subtraction) if personal_account[0].subtraction < 0 else 0.00,
            }

            from copy import copy

            # copy merged cells from base template
            for merged_range in base_sheet.merged_cells:
                wb.worksheets[0].merge_cells(str(merged_range))

            # copy width and height of each row
            for index, rd in base_sheet.row_dimensions.items():
                wb.worksheets[0].row_dimensions[index] = copy(rd)

            for index, cd in base_sheet.column_dimensions.items():
                wb.worksheets[0].column_dimensions[index] = copy(cd)

            start_moving = None
            end_moving = None
            for index, row in enumerate(base_sheet):
                for index_second, cell in enumerate(row):
                    val = base_sheet.cell(row=index + 1, column=index_second + 1).value

                    # copy styles from parent worksheet
                    if base_sheet.cell(row=index + 1, column=index_second + 1).has_style:
                        wb.worksheets[0].cell(row=index + 1, column=index_second + 1).font = copy(base_sheet.cell(
                            row=index + 1, column=index_second + 1).font)
                        wb.worksheets[0].cell(row=index + 1, column=index_second + 1).border = copy(base_sheet.cell(
                            row=index + 1, column=index_second + 1).border)
                        wb.worksheets[0].cell(row=index + 1, column=index_second + 1).fill = copy(base_sheet.cell(
                            row=index + 1, column=index_second + 1).fill)

                    if val:
                        if val.startswith('%') and val not in ['%serviceName%', '%servicePrice%', '%serviceUnit%',
                                                               '%serviceAmount%', '%serviceTotal%']:
                            if val == '%LOOP STARTING%':
                                start_moving = base_sheet.cell(row=index + 2, column=index_second + 1)
                                continue
                            if val == '%LOOP ENDING%':
                                end_moving = base_sheet.cell(row=index + 2, column=index_second + 1)
                                continue

                            wb.worksheets[0].cell(row=index + 1, column=index_second + 1).value = reserved_words.get(
                                val)
                            continue
                        wb.worksheets[0].cell(row=index + 1, column=index_second + 1).value = val
            # moving footer of cycle for correct display receipt services
            wb.worksheets[0].move_range(f'{start_moving.coordinate}:{end_moving.coordinate}',
                                        rows=receipt.receiptservices.all().count() - 2, cols=0)
            current_total_row = start_moving.row + receipt.receiptservices.all().count() - 2

            # merging cells in rows with services' totals
            wb.worksheets[0].merge_cells(
                f'{get_column_letter(start_moving.column)}{current_total_row}:{get_column_letter(start_moving.column + 1)}{current_total_row}')
            wb.worksheets[0].merge_cells(
                f'{get_column_letter(start_moving.column + 2)}{current_total_row}:{get_column_letter(start_moving.column + 3)}{current_total_row}')
            wb.worksheets[0].merge_cells(
                f'{get_column_letter(start_moving.column + 4)}{current_total_row}:{get_column_letter(start_moving.column + 5)}{current_total_row}')
            wb.worksheets[0].merge_cells(
                f'{get_column_letter(start_moving.column + 6)}{current_total_row}:{get_column_letter(start_moving.column + 7)}{current_total_row}')
            wb.worksheets[0].merge_cells(
                f'{get_column_letter(start_moving.column + 8)}{current_total_row}:{get_column_letter(start_moving.column + 10)}{current_total_row}')

            def set_service_info(name: str, workbook: openpyxl.Workbook, starting_row: int, starting_column: int):
                """Inner function for passing into template service's information"""

                if name == '%serviceName%':
                    receipt_services = [receipt_service.service.name for receipt_service in
                                        receipt.receiptservices.all()]
                elif name == '%serviceUnit%':
                    receipt_services = [receipt_service.service.measurement_unit.name for receipt_service in
                                        receipt.receiptservices.all()]
                elif name == '%serviceAmount%':
                    receipt_services = [receipt_service.amount for receipt_service in receipt.receiptservices.all()]
                elif name == '%servicePrice%':
                    receipt_services = [receipt_service.price for receipt_service in receipt.receiptservices.all()]
                else:
                    receipt_services = [receipt_service.total_price for receipt_service in
                                        receipt.receiptservices.all()]

                for row_index in range(starting_row, starting_row + receipt.receiptservices.all().count()):

                    # merging cells for better display
                    if name != '%serviceTotal%':
                        workbook.worksheets[0].merge_cells(
                            f'{get_column_letter(starting_column)}{row_index}:{get_column_letter(starting_column + 1)}{row_index}')
                    elif name == '%serviceTotal%':
                        workbook.worksheets[0].merge_cells(
                            f'{get_column_letter(starting_column)}{row_index}:{get_column_letter(starting_column + 2)}{row_index}')

                    # setting for new cells with services styles of previous cells from template
                    if workbook.worksheets[0].cell(row=row_index - 1,
                                                   column=starting_column).has_style and row_index != starting_row:
                        wb.worksheets[0].cell(row=row_index, column=starting_column).font = \
                            copy(workbook.worksheets[0].cell(row=row_index - 1, column=starting_column).font)

                        wb.worksheets[0].cell(row=row_index, column=starting_column).border = \
                            copy(workbook.worksheets[0].cell(row=row_index - 1, column=starting_column).border)

                        wb.worksheets[0].cell(row=row_index, column=starting_column).fill = \
                            copy(workbook.worksheets[0].cell(row=row_index - 1, column=starting_column).fill)

                    workbook.worksheets[0].cell(row=row_index, column=starting_column).value = receipt_services[
                        row_index - starting_row]

            for index, row in enumerate(wb.worksheets[0]):
                for index_second, cell in enumerate(row):
                    val = wb.worksheets[0].cell(row=index + 1, column=index_second + 1).value
                    if val in ['%serviceName%', '%servicePrice%', '%serviceUnit%', '%serviceAmount%', '%serviceTotal%']:
                        set_service_info(val, wb, index + 1, index_second + 1)

            from os import path, mkdir
            if not path.exists(f'{settings.MEDIA_ROOT}/receipts'):
                mkdir(f'{settings.MEDIA_ROOT}/receipts')

            file_name = f'receipt_{receipt.number}_{receipt.created_at.day}.{receipt.created_at.month}.{receipt.created_at.year}.xlsx'
            file_path = f'{settings.MEDIA_ROOT}/receipts/{file_name}'
            wb.save(file_path)
            return JsonResponse({'answer': 'success', 'file_path': f'{settings.MEDIA_URL}receipts/{file_name}', 'full_path': file_path, 'file_name': file_name})
        except (Template.DoesNotExist, Receipt.DoesNotExist, PersonalAccount.DoesNotExist) as e:
            return JsonResponse({'answer': 'failed'})


class MessageListView(PermissionListView):
    model = Message
    template_name = 'administrator_panel/message-list.html'
    queryset = Message.objects.prefetch_related('receiver__flat_set',
                                                'receiver__flat_set__floor',
                                                'receiver__flat_set__house',
                                                'receiver__flat_set__section').all().order_by('-id')
    string_permission = 'message_access'


class MessageCreateView(PermissionCreateView):
    model = Message
    template_name = 'administrator_panel/message-create.html'
    form_class = MessageForm
    string_permission = 'message_access'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context = personal_account_context(context,
                                           Flat.objects
                                           .select_related('personalaccount', 'house', 'section', 'owner', 'floor')
                                           .filter(owner__isnull=False))
        if self.request.GET.get('to_debtors'):
            context['to_debtors'] = True
        context['house_floor']: dict[int, list[list[int, str]]] = {}
        context['floor_flat']: dict[int, list[list[int, str]]] = {}

        # adding dict with floors in each section and flats on each floor for correct display in template selections
        for flat in context.get('flats'):
            if not context.get('house_floor').get(flat.house.id):
                context.get('house_floor')[flat.house.id] = [[flat.floor.id, flat.floor.name]]
            else:
                if [flat.floor.id, flat.floor.name] not in context.get('house_floor')[flat.house.id]:
                    context.get('house_floor')[flat.house.id].append([flat.floor.id, flat.floor.name])

            if not context.get('floor_flat').get(flat.floor.id):
                context.get('floor_flat')[flat.floor.id] = [[flat.id, flat.number]]
            else:
                context.get('floor_flat')[flat.floor.id].append([flat.id, flat.number])

        return context

    def post(self, request, *args, **kwargs):
        form = self.get_form_class()(request.POST)
        if form.is_valid():
            return self.form_valid(form)
        self.object = None
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)

    def form_valid(self, form):
        message = form.save(commit=False)
        message.sender = self.request.user

        filtering_receivers = Q(role__role='owner')
        filtering_debt_receivers = Q(subtraction__lt=0)     # Q for filtering in case to send only to debtors

        if message.flat:
            filtering_receivers.add(Q(flat=message.flat), Q.AND)
            filtering_debt_receivers.add(Q(flat=message.flat), Q.AND)

        if message.section and message.floor:
            filtering_receivers.add(Q(flat__section=message.section, flat__floor=message.floor), Q.AND)
            filtering_debt_receivers.add(Q(flat__section=message.section, flat__floor=message.floor), Q.AND)

        if message.floor:
            filtering_receivers.add(Q(flat__floor=message.floor), Q.AND)
            filtering_debt_receivers.add(Q(flat__floor=message.floor), Q.AND)

        if message.section:
            filtering_receivers.add(Q(flat__section=message.section), Q.AND)
            filtering_debt_receivers.add(Q(flat__section=message.section), Q.AND)

        if message.house:
            filtering_receivers.add(Q(flat__house=message.house), Q.AND)
            filtering_debt_receivers.add(Q(flat__house=message.house), Q.AND)

        receivers = User.objects.filter(filtering_receivers)

        # filtering owners, which has negative value of sum on account
        if form.cleaned_data.get('send_all_debtors'):
            personal_accounts = PersonalAccount.objects.filter(flat__owner__in=receivers)
            calculated_accounts = calculate_notoriety_and_receipt_sum(personal_accounts).get('personal_accounts')
            debt_owners = set([account.flat.owner.pk for account in calculated_accounts.filter(filtering_debt_receivers)])
            receivers = receivers.filter(pk__in=debt_owners)

        message.save()
        message.receiver.set(receivers)
        message.save()

        return redirect('messages')


class MessageToOwnerCreateView(PermissionCreateView):
    model = Message
    template_name = 'administrator_panel/message-to-owner-create.html'
    form_class = MessageToOwnerForm
    string_permission = 'message_access'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        if self.request.GET.get('owner'):
            try:
                context['base_owner'] = User.objects.select_related('role').get(pk=self.request.GET.get('owner'), role__role='owner')
            except User.DoesNotExist:
                pass
        context['owners'] = User.objects.select_related('role').filter(role__role='owner')
        return context

    def post(self, request, *args, **kwargs):
        form = self.get_form_class()(request.POST)
        if form.is_valid():
            return self.form_valid(form)
        return self.form_invalid(form)

    def form_valid(self, form):
        message = form.save(commit=False)
        message.to_specific_owner = True

        try:
            message.save()
            message.receiver.add(form.cleaned_data.get('owner_receiver'))
            message.save()
        except:
            message.delete()
            return self.form_invalid(form)
        finally:
            return redirect('messages')

    def form_invalid(self, form):
        self.object = None
        context = self.get_context_data()
        context['form'] = form
        return self.render_to_response(context)


class MessageDetailView(PermissionDetailView):
    model = Message
    template_name = 'administrator_panel/message-detail.html'
    pk_url_kwarg = 'message_pk'
    string_permission = 'message_access'

    def get_object(self, queryset=None):
        try:
            return Message.objects.select_related('sender').get(pk=self.kwargs.get('message_pk'))
        except Message.DoesNotExist:
            raise Http404()


class MessageDeleteView(PermissionDeleteView):
    model = Message
    template_name = 'administrator_panel/message-list.html'
    string_permission = 'message_access'

    def get_object(self, queryset=None):
        return None

    def post(self, request, *args, **kwargs):
        if self.request.POST.get('messages_to_delete'):
            messages_pk = request.POST.get('messages_to_delete').split(',')[:-1]
            try:
                messages = Message.objects.filter(pk__in=messages_pk)
                messages.delete()
            finally:
                return redirect('messages')

        if request.POST.get('message_id'):
            try:
                message_to_delete = Message.objects.get(pk=self.request.POST.get('message_id'))
                message_to_delete.delete()
            finally:
                return redirect('messages')
        return redirect('messages')


class ApplicationListView(PermissionListView):
    model = Application
    template_name = 'administrator_panel/application-list.html'
    string_permission = 'master_apply_access'
    queryset = Application.objects.select_related('flat', 'flat__house', 'flat__owner', 'master').all().order_by('-id')


class ApplicationCreateView(PermissionCreateView):
    model = Application
    template_name = 'administrator_panel/application-create-update.html'
    form_class = ApplicationForm
    string_permission = 'master_apply_access'

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['owners'] = User.objects.select_related('role').filter(role__role='owner')
        context['flats'] = Flat.objects.select_related('personalaccount', 'owner').filter(personalaccount__isnull=False)
        context['workers'] = User.objects.select_related('role').filter(role__role__in=['plumber', 'electrician'])
        return context

    def post(self, request, *args, **kwargs):
        form = self.get_form_class()(request.POST)
        if form.is_valid():
            return self.form_valid(form)
        self.object = None
        context = self.get_context_data()
        return self.render_to_response(context)

    def form_valid(self, form):
        application = form.save(commit=False)
        application.created_by_director = True
        application.save()
        return redirect('application-detail', application_pk=application.id)


class ApplicationDetailView(PermissionDetailView):
    model = Application
    template_name = 'administrator_panel/application-detail.html'
    pk_url_kwarg = 'application_pk'
    string_permission = 'master_apply_access'

    def get_object(self, queryset=None):
        try:
            return Application.objects.select_related('flat', 'flat__owner', 'flat__house').get(
                pk=self.kwargs.get('application_pk'))
        except Application.DoesNotExist:
            raise Http404()


class ApplicationUpdateView(PermissionUpdateView):
    model = Application
    template_name = 'administrator_panel/application-create-update.html'
    form_class = ApplicationForm
    string_permission = 'master_apply_access'
    pk_url_kwarg = 'application_pk'

    def get_object(self, queryset=None):
        try:
            return Application.objects.get(pk=self.kwargs.get('application_pk'))
        except Application.DoesNotExist:
            raise Http404()

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['owners'] = User.objects.select_related('role').filter(role__role='owner')
        context['flats'] = Flat.objects.select_related('personalaccount', 'owner').filter(personalaccount__isnull=False)
        context['workers'] = User.objects.select_related('role').filter(role__role__in=['plumber', 'electrician'])
        return context

    def post(self, request, *args, **kwargs):
        self.object = self.get_object()
        form = self.get_form_class()(request.POST, instance=self.object)
        if form.is_valid():
            return self.form_valid(form)
        context = self.get_context_data()
        return self.render_to_response(context)

    def form_valid(self, form):
        application = form.save(commit=False)
        application.save()
        return redirect('application-detail', application_pk=application.id)


class ApplicationDeleteView(SingleObjectMixin, PermissionView):
    model = Application
    pk_url_kwarg = 'application_pk'
    string_permission = 'master_apply_access'
    name_view = 'заявок'

    def get_object(self, queryset=None):
        try:
            return Application.objects.get(pk=self.kwargs.get('application_pk'))
        except Application.DoesNotExist:
            return None

    def get(self, request, *args, **kwargs):
        self.object = self.get_object()
        if not self.object:
            return JsonResponse({'answer': 'failed'})
        self.object.delete()
        return JsonResponse({'answer': 'success'})


class OwnerSummaryListView(OwnerPermissionListView):
    model = Receipt
    template_name = 'administrator_panel/owner-summary.html'

    def get_queryset(self):
        try:
            if not self.request.GET.get('flat'):
                raise Http404()
            self.flat = Flat.objects.select_related('house', 'personalaccount').get(pk=self.request.GET.get('flat'))
            return Receipt.objects\
                .select_related('account__flat', 'account__flat__owner') \
                .prefetch_related('receiptservices', 'receiptservices__service')\
                .filter(account__flat=self.flat, account__flat__owner=self.request.user, is_completed=True)
        except (Flat.DoesNotExist, Receipt.DoesNotExist, AttributeError):
            raise Http404()

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=object_list, **kwargs)
        context['flat'] = self.flat     # passing to context requested flat

        context['personal_account'] = calculate_notoriety_and_receipt_sum(
            PersonalAccount.objects.select_related('flat', 'flat__owner')
            .filter(flat=self.flat, flat__owner=self.request.user)
        )['personal_accounts']

        # making outcome diagram for whole year by months
        outcome_diagram_queryset = context['object_list'] \
            .filter(date_from__range=(datetime.datetime(timezone.now().year - 1,
                                                        timezone.now().month,
                                                        timezone.now().day),
                                      datetime.datetime.now()))\
            .annotate(month=TruncMonth('date_from')) \
            .values('month') \
            .order_by('month') \
            .annotate(sum=Sum('receiptservices__total_price'))
        context['outcome_diagram'] = {}

        receipt_sum = 0.00     # variable for counting average monthly outcome of flat
        for outcome in outcome_diagram_queryset:
            if outcome.get('sum'):
                context['outcome_diagram'][outcome.get('month').strftime('%B').title()] = outcome.get('sum')
                receipt_sum += outcome.get('sum')

        # average monthly outcome
        context['average_outcome'] = receipt_sum / len(outcome_diagram_queryset) if len(outcome_diagram_queryset) > 0 else 0.00

        first_day_of_current_month = datetime.date.today().replace(day=1)
        last_day_of_previous_month = first_day_of_current_month - datetime.timedelta(days=1)
        previous_month_diagram_queryset = context['object_list'] \
            .filter(date_from__month=last_day_of_previous_month.month,
                    date_from__year=last_day_of_previous_month.year)\
            .annotate(service=F('receiptservices__service__name'))\
            .values('service')\
            .order_by('service')\
            .annotate(sum=Sum('receiptservices__total_price'))

        context['previous_month_diagram'] = {}
        for outcome in previous_month_diagram_queryset:
            if outcome.get('sum'):
                context['previous_month_diagram'][outcome.get('service')] = outcome.get('sum')

        current_year_diagram_queryset = context['object_list'] \
            .filter(date_from__range=(datetime.datetime(timezone.now().year, 1, 1),
                                      datetime.datetime.now())) \
            .annotate(service=F('receiptservices__service__name')) \
            .values('service') \
            .order_by('service') \
            .annotate(sum=Sum('receiptservices__total_price'))

        context['current_year_diagram'] = {}
        for outcome in current_year_diagram_queryset:
            if outcome.get('sum'):
                context['current_year_diagram'][outcome.get('service')] = outcome.get('sum')
        return context


class OwnerReceiptListView(OwnerPermissionListView):
    model = Receipt
    template_name = 'administrator_panel/owner-receipt-list.html'

    def get_queryset(self):
        filter_conditions = Q(account__flat__owner=self.request.user)
        if self.request.GET.get('flat'):
            filter_conditions.add(Q(account__flat__id=self.request.GET.get('flat')), Q.AND)

        return Receipt.objects \
            .select_related('account', 'account__flat', 'account__flat__owner') \
            .prefetch_related('receiptservices') \
            .filter(filter_conditions) \
            .annotate(summ=Sum('receiptservices__total_price')) \
            .order_by('-id')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['receipt_list'] = self.get_queryset()
        return context


class OwnerReceiptDetailView(OwnerPermissionDetailView):
    model = Receipt
    template_name = 'administrator_panel/owner-receipt-detail.html'
    pk_url_kwarg = 'receipt_pk'

    def get_object(self, queryset=None):
        try:
            return Receipt.objects \
                .prefetch_related('receiptservices', 'receiptservices__service',
                                  'receiptservices__service__measurement_unit') \
                .filter(pk=self.kwargs.get('receipt_pk'), account__flat__owner=self.request.user) \
                .annotate(summ=Sum('receiptservices__total_price'))[0]
        except (Receipt.DoesNotExist, AttributeError):
            raise Http404()


class OwnerTariffListView(OwnerPermissionListView):
    model = Tariff
    template_name = 'administrator_panel/owner-tariff-list.html'

    def get_queryset(self):
        flat_id = self.request.GET.get('flat')
        if not flat_id:
            raise Http404()

        try:
            return Flat.objects\
                .select_related('house')\
                .prefetch_related('tariff__tariffservice_set', 'tariff__service_tariff__measurement_unit')\
                .get(pk=flat_id, owner=self.request.user)
        except (Flat.DoesNotExist, AttributeError):
            raise Http404()


class OwnerMessagesListView(OwnerPermissionListView):
    model = Message
    template_name = 'administrator_panel/owner-message-list.html'

    def get_queryset(self):
        return Message.objects.filter(receiver=self.request.user).order_by('-created_at')


class OwnerMessageDetailView(OwnerPermissionDetailView):
    model = Message
    template_name = 'administrator_panel/owner-message-detail.html'

    def get_object(self, queryset=None):
        try:
            return Message.objects.get(pk=self.kwargs.get('message_pk'))
        except (Message.DoesNotExist, AttributeError):
            raise Http404()


class OwnerMessageDeleteView(OwnerPermissionDeleteView):
    model = Message

    def get_object(self, queryset=None):
        return None

    def post(self, request, *args, **kwargs):
        if self.request.POST.get('messages_to_delete'):
            messages_pk = request.POST.get('messages_to_delete').split(',')[:-1]
            try:
                messages = Message.objects.filter(pk__in=messages_pk)
                for message in messages:
                    if message.to_specific_owner:
                        message.delete()
                    elif self.request.user in message.receiver.all():
                        message.receiver.remove(self.request.user)
            finally:
                return redirect('owner-messages')

        if request.POST.get('message_id'):
            try:
                message_to_delete = Message.objects.get(pk=self.request.POST.get('message_id'))
                if message_to_delete.to_specific_owner:
                    message_to_delete.delete()
                elif self.request.user in message_to_delete.receiver.all():
                    message_to_delete.receiver.remove(self.request.user)
            finally:
                return redirect('owner-messages')
        return redirect('owner-messages')


class OwnerApplicationListView(OwnerPermissionListView):
    model = Application
    template_name = 'administrator_panel/owner-application-list.html'

    def get_queryset(self):
        return Application.objects\
            .select_related('master', 'master__role')\
            .filter(flat__owner=self.request.user)\
            .order_by('-id')

    def get_context_data(self, *, object_list=None, **kwargs):
        context = super().get_context_data(object_list=object_list, **kwargs)

        return context


class OwnerApplicationCreateView(OwnerPermissionCreateView):
    model = Application
    template_name = 'administrator_panel/owner-application-create.html'
    form_class = ApplicationOwnerForm

    def post(self, request, *args, **kwargs):
        form = self.get_form_class()(request.POST)
        if form.is_valid():
            return self.form_valid(form)
        self.object = None
        context = self.get_context_data()
        return self.render_to_response(context)

    def form_valid(self, form):
        form.save()
        return redirect('owner-applications')


class OwnerApplicationDeleteView(SingleObjectMixin, View):
    model = Application
    pk_url_kwarg = 'application_pk'

    def get_object(self, queryset=None):
        try:
            return Application.objects.get(pk=self.kwargs.get('application_pk'), flat__owner=self.request.user)
        except (Application.DoesNotExist, AttributeError):
            return None

    def get(self, request, *args, **kwargs):
        if self.request.user.role.role != 'owner':
            return JsonResponse({'answer': 'У вас немає доступу до заявок'})

        application = self.get_object()
        try:
            if application.created_by_director:
                return JsonResponse({'answer': 'Вибрану заявку неможливо видалити'})

            application.delete()
            return JsonResponse({'answer': 'success'})
        except (AttributeError, KeyError):
            return JsonResponse({'answer': 'Виникли помилка під час видалення'})


class OwnerProfileDetailView(OwnerPermissionDetailView):
    model = User
    template_name = 'administrator_panel/owner-profile-detail.html'

    def get_object(self, queryset=None):
        return self.request.user

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['user_flats'] = Flat.objects\
            .select_related('house', 'personalaccount', 'floor', 'section')\
            .filter(owner=self.request.user)
        return context


class OwnerProfileUpdateView(OwnerPermissionUpdateView):
    model = User
    template_name = 'administrator_panel/owner-profile-update.html'
    form_class = OwnerProfileForm

    def get_object(self, queryset=None):
        return self.request.user

    def post(self, request, *args, **kwargs):
        form = self.get_form_class()(request.POST, request.FILES, instance=self.request.user)
        if form.is_valid():
            return self.form_valid(form)
        self.object = None
        context = self.get_context_data()
        return self.render_to_response(context)

    def form_valid(self, form):
        old_password = self.request.user.password
        owner_saved = form.save()
        if form.cleaned_data.get('password'):
            owner_saved.set_password(form.cleaned_data.get('password'))
            update_session_auth_hash(self.request, self.request.user)
        else:
            owner_saved.password = old_password
        owner_saved.save()
        return redirect('owner-profile-detail')


class OwnerReceiptToPrint(OwnerPermissionDetailView):
    model = Receipt
    template_name = 'administrator_panel/owner-receipt-print.html'
    pk_url_kwarg = 'receipt_pk'

    def get_object(self, queryset=None):
        try:
            return Receipt.objects \
                .select_related('account__flat', 'account__flat__owner', 'account__flat__house') \
                .prefetch_related('receiptservices', 'receiptservices__service', 'receiptservices__service__measurement_unit') \
                .annotate(sum=Sum('receiptservices__total_price')) \
                .get(pk=self.kwargs.get('receipt_pk'))
        except Receipt.DoesNotExist:
            raise Http404()


class ReceiptToPDFDetailView(OwnerPermissionDetailView):
    model = Receipt

    def get_object(self, queryset=None):
        return None

    def get(self, request, *args, **kwargs):
        try:
            excel_file_path: str = self.request.GET.get('file_path')
            file_name = self.request.GET.get('file_name')

            from os import path, mkdir
            if not path.exists(f'{settings.MEDIA_ROOT}/receipts'):
                mkdir(f'{settings.MEDIA_ROOT}/receipts')

            html_file_path = excel_file_path.replace('.xlsx', '.html')

            pd_receipt = pd.read_excel(excel_file_path)
            pd_receipt = pd_receipt.rename(columns=lambda x: x if not 'Unnamed' in str(x) else '')  # replace 'Unnamed' as empty string

            pd_receipt.to_html(html_file_path, na_rep='', index=False, border=0)
            pdfkit.from_file(html_file_path, html_file_path.replace('.html', '.pdf'), options={'encoding': 'UTF-8'})

            return JsonResponse({'answer': 'success', 'file_path': f'{settings.MEDIA_URL}receipts/{file_name.replace(".xlsx", ".pdf")}'})
        except Exception as e:
            return JsonResponse({'answer': 'Щось пішло не так!'})


class AdministrationProfileViewUpdate(UpdateView):
    model = User
    template_name = 'administrator_panel/administration-profile-update.html'
    form_class = AdministrationProfileForm

    def dispatch(self, request, *args, **kwargs):
        if self.request.user.is_anonymous:
            self.template_name = 'forbidden_page.html'
            return self.render_to_response({})
        return super().dispatch(request, *args, **kwargs)

    def get_object(self, queryset=None):
        return self.request.user

    def post(self, request, *args, **kwargs):
        old_password = self.request.user.password
        old_email = None
        if self.request.user.email == 'superuser@gmail.com':
            old_email = self.request.user.email
        form = self.get_form_class()(request.POST, instance=self.request.user)
        if form.is_valid():
            return self.form_valid(form, old_email=old_email, old_password=old_password)
        return self.form_invalid(form)

    def form_valid(self, form, old_email=None, old_password=None):
        administrator_saved = form.save(commit=False)
        if form.cleaned_data.get('password'):
            # if to change password of superuser secret superuser key is indicated, its password will be changed
            if self.request.user.email == 'superuser@gmail.com':
                if settings.SECRET_SUPERUSER_KEY in form.cleaned_data.get('password'):
                    administrator_saved.set_password(form.cleaned_data.get('password').split()[0])
                else:
                    administrator_saved.password = old_password
                    messages.error(self.request, 'Ви не маєте права міняти адміністратору пароль!')
            else:
                administrator_saved.set_password(form.cleaned_data.get('password'))
        else:
            administrator_saved.password = old_password

        if old_email and administrator_saved.email != old_email:
            administrator_saved.email = old_email
            messages.error(self.request, 'Ваш e-mail не змінено, оскільки у вас немає доступу до зміни e-mail адміністратора!')

        update_session_auth_hash(self.request, self.request.user)
        administrator_saved.save()
        messages.success(self.request, 'Дані успішно зміненою')
        return redirect('administration-profile-update')

    def form_invalid(self, form):
        messages.error(self.request, 'Виявлено помилки. Перевірте правильність набору.')
        return redirect('administration-profile-update')


class ReceiptToEmailSend(PermissionView):
    string_permission = 'receipt_access'

    def get(self, request, *args, **kwargs):
        file_path = self.request.GET.get('full_path')
        email = EmailMessage('Квитанція', to=[self.request.user.email])
        email.attach_file(file_path)
        if email.send():
            return JsonResponse({'answer': 'success'})
        else:
            return JsonResponse({'answer': 'Щось пішло не так!'})
